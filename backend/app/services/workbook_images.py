from __future__ import annotations

import re
from collections.abc import Iterable, Mapping
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook

WorkbookSource = bytes | bytearray | str | Path | BinaryIO
IMAGE_HEADER = re.compile(r"^image[ _-]?(\d+)$", re.IGNORECASE)


def _key(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _alias_lookup(
    aliases: Mapping[str, str | Iterable[str]] | None,
) -> dict[str, str]:
    lookup: dict[str, str] = {}
    for canonical, values in (aliases or {}).items():
        if isinstance(values, str):
            lookup[_key(canonical)] = str(canonical)
            lookup[_key(values)] = str(canonical)
        else:
            lookup[_key(canonical)] = str(canonical)
            lookup.update({_key(value): str(canonical) for value in values})
    return lookup


def canonicalize_header(
    header: object, aliases: Mapping[str, str | Iterable[str]] | None = None
) -> str:
    text = str(header or "").strip()
    image_match = IMAGE_HEADER.fullmatch(text)
    if image_match:
        return f"image_{int(image_match.group(1))}"
    return _alias_lookup(aliases).get(_key(text), text)


def parse_image_workbook(
    source: WorkbookSource,
    aliases: Mapping[str, str | Iterable[str]] | None = None,
    *,
    max_image_columns: int = 50,
) -> dict[str, object]:
    """Parse the active sheet while retaining raw and canonical cell values."""
    stream: object = BytesIO(source) if isinstance(source, (bytes, bytearray)) else source
    workbook = load_workbook(stream, read_only=True, data_only=False)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    raw_headers = list(next(iterator, ()))
    while raw_headers and raw_headers[-1] is None:
        raw_headers.pop()
    if not raw_headers:
        raise ValueError("workbook has no header row")

    headers = [canonicalize_header(header, aliases) for header in raw_headers]
    folded = [_key(header) for header in headers]
    if len(folded) != len(set(folded)):
        raise ValueError("workbook contains duplicate headers after alias mapping")

    image_columns = [header for header in headers if IMAGE_HEADER.fullmatch(header)]
    if len(image_columns) > max_image_columns:
        raise ValueError(f"workbook has more than {max_image_columns} image columns")

    rows: list[dict[str, object]] = []
    for row_number, cells in enumerate(iterator, start=2):
        values = list(cells[: len(headers)])
        values.extend([None] * (len(headers) - len(values)))
        if not any(value is not None and value != "" for value in values):
            continue
        canonical = dict(zip(headers, values, strict=True))
        raw = {str(header): value for header, value in zip(raw_headers, values, strict=True)}
        rows.append(
            {
                "row_number": row_number,
                "values": canonical,
                "raw": raw,
                "images": [canonical[column] for column in image_columns],
            }
        )

    workbook.close()
    return {
        "sheet": sheet.title,
        "raw_headers": raw_headers,
        "headers": headers,
        "image_columns": image_columns,
        "rows": rows,
    }
